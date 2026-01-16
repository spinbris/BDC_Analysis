"""Export analysis results to Excel"""

from pathlib import Path
from typing import Dict, Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


def to_excel(
    results: Dict[str, Any],
    output_path: Path,
    bdc_info: Dict[int, Dict[str, str]] = None
) -> None:
    """Export analysis results to multi-tab Excel workbook

    Args:
        results: Dictionary with analysis results:
            - 'summary': BDC summary DataFrame
            - 'common_holdings': Common holdings DataFrame
            - 'overlap_matrix': BDC × BDC overlap matrix
            - 'holdings_matrix': Optional full holdings matrix
        output_path: Path to write Excel file
        bdc_info: Optional dict mapping CIK to ticker/name
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    print(f"\nExporting results to Excel...")
    print(f"  Output: {output_path}")

    # Create Excel writer
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:

        # Tab 1: Summary
        if 'summary' in results:
            summary = results['summary'].copy()

            # Add ticker/name if available
            if bdc_info:
                summary['ticker'] = summary['bdc_cik'].map(
                    lambda x: bdc_info.get(x, {}).get('ticker', '')
                )
                summary['bdc_name'] = summary['bdc_cik'].map(
                    lambda x: bdc_info.get(x, {}).get('name', '')
                )
                # Reorder columns
                cols = ['ticker', 'bdc_name', 'bdc_cik', 'total_companies',
                        'shared_companies', 'pct_shared']
                summary = summary[[c for c in cols if c in summary.columns]]

            summary.to_excel(writer, sheet_name='BDC Summary', index=False)
            print(f"  ✓ Tab: BDC Summary ({len(summary)} rows)")

        # Tab 2: Common Holdings
        if 'common_holdings' in results:
            common = results['common_holdings'].copy()

            # Format holders list as string
            if 'holders' in common.columns:
                common['holders'] = common['holders'].apply(
                    lambda x: ', '.join(map(str, x)) if isinstance(x, list) else str(x)
                )

            common.to_excel(writer, sheet_name='Common Holdings', index=False)
            print(f"  ✓ Tab: Common Holdings ({len(common)} rows)")

        # Tab 3: BDC Overlap Matrix
        if 'overlap_matrix' in results:
            overlap = results['overlap_matrix'].copy()

            # Add ticker labels if available
            if bdc_info:
                overlap.index = overlap.index.map(
                    lambda x: f"{bdc_info.get(x, {}).get('ticker', x)} ({x})"
                )
                overlap.columns = overlap.columns.map(
                    lambda x: f"{bdc_info.get(x, {}).get('ticker', x)} ({x})"
                )

            overlap.to_excel(writer, sheet_name='BDC Overlap Matrix')
            print(f"  ✓ Tab: BDC Overlap Matrix ({overlap.shape[0]}×{overlap.shape[1]})")

        # Tab 4: Methodology
        methodology_df = pd.DataFrame({
            'Section': [
                'Data Source',
                'Data Date',
                'BDCs Analyzed',
                'Matching Method',
                'Common Holdings Definition',
                'Notes'
            ],
            'Description': [
                'SEC BDC Data Sets (Schedule of Investments)',
                pd.Timestamp.now().strftime('%Y-%m-%d'),
                f"{len(results.get('summary', []))} BDCs",
                'Exact company name match (Phase 1)',
                'Companies held by 2+ BDCs',
                'Phase 1 prototype - exact matching only. Entity resolution in Phase 2.'
            ]
        })
        methodology_df.to_excel(writer, sheet_name='Methodology', index=False)
        print(f"  ✓ Tab: Methodology")

    # Apply formatting
    _apply_excel_formatting(output_path)

    print(f"✓ Export complete: {output_path}")


def _apply_excel_formatting(file_path: Path) -> None:
    """Apply formatting to Excel workbook

    Args:
        file_path: Path to Excel file to format
    """
    wb = load_workbook(file_path)

    # Header formatting
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Format headers (row 1)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(file_path)


if __name__ == "__main__":
    # Test export
    from pathlib import Path
    from load import load_submissions, load_soi, filter_top_bdcs, TOP_10_BDCS
    from overlap import (
        build_holdings_matrix,
        find_common_holdings,
        calculate_bdc_overlap_matrix,
        get_overlap_summary
    )

    data_dir = Path(__file__).parent.parent / "data" / "raw" / "bdc-data-sets"

    # Load data
    print("Loading data...")
    sub_df = load_submissions(data_dir)
    soi_df = load_soi(data_dir)
    filtered_soi = filter_top_bdcs(soi_df, sub_df)

    # Find company name column
    name_cols = [c for c in filtered_soi.columns if 'name' in c.lower()]
    company_col = name_cols[0]

    # Run analysis
    print("\nRunning analysis...")
    matrix = build_holdings_matrix(filtered_soi, company_col, bdc_col='cik')
    common = find_common_holdings(matrix, min_holders=2)
    overlap_matrix = calculate_bdc_overlap_matrix(matrix)
    summary = get_overlap_summary(matrix)

    # Export
    results = {
        'summary': summary,
        'common_holdings': common,
        'overlap_matrix': overlap_matrix
    }

    output_path = Path(__file__).parent.parent / "outputs" / "bdc_overlap_phase1.xlsx"
    to_excel(results, output_path, bdc_info=TOP_10_BDCS)

    print(f"\n✓ Test complete - check {output_path}")
