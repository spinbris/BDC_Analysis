"""
Generate Excel report with industry analysis
Run: python -m src.generate_industry_report
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

print("=" * 70)
print("GENERATING BDC INDUSTRY ANALYSIS REPORT")
print("=" * 70)

# Load data
print("\nLoading analysis results...")
df = pd.read_csv('outputs/all_bdc_holdings_with_industry.csv')
industry_counts = pd.read_csv('outputs/industry_company_counts_by_bdc.csv', index_col=0)
industry_fv = pd.read_csv('outputs/industry_fair_value_by_bdc.csv', index_col=0)
overlap_analysis = pd.read_csv('outputs/industry_overlap_analysis.csv')

# Filter to BDCs with descriptions
bdcs_with_data = ['Ares Capital Corporation', 'Main Street Capital Corporation']
df_analyzed = df[df['bdc_name'].isin(bdcs_with_data)]

print(f"✓ Loaded {len(df):,} total holdings")
print(f"✓ Analyzing {len(df_analyzed):,} holdings from 2 BDCs with business descriptions")

# Create Excel report
output_path = Path('outputs/bdc_industry_analysis.xlsx')

print("\n" + "=" * 70)
print("CREATING EXCEL WORKBOOK")
print("=" * 70)

with pd.ExcelWriter(output_path, engine='openpyxl') as writer:

    # Tab 1: Executive Summary
    exec_summary = pd.DataFrame({
        'Metric': [
            'BDCs Analyzed',
            'Total Holdings Analyzed',
            'Unique Portfolio Companies',
            'Total Fair Value ($B)',
            '',
            'Industry Sectors Identified',
            'Companies with Classification',
            'Coverage Rate',
            '',
            'Top Industry (ARCC)',
            'Top Industry (MAIN)',
            '',
            'Companies in Software/Tech',
            'Software/Tech Fair Value ($B)',
            '',
            'Company-Level Overlaps',
            'Industry Concentration (ARCC HHI)',
            'Industry Concentration (MAIN HHI)',
        ],
        'Value': [
            2,
            len(df_analyzed),
            len(df_analyzed.groupby(['bdc_name', 'company_name']).size()),
            df_analyzed['fair_value'].sum() / 1e9,
            '',
            len(df_analyzed['industry_sector'].unique()) - 1,  # Exclude Unknown
            df_analyzed[df_analyzed['industry_sector'] != 'Unknown'].shape[0],
            f"{df_analyzed[df_analyzed['industry_sector'] != 'Unknown'].shape[0] / len(df_analyzed) * 100:.1f}%",
            '',
            'Software/Technology (36.2%)',
            'Other/Mixed (32.4%)',
            '',
            140,
            13.03,
            '',
            1,
            '2576 (High)',
            '1702 (Moderate)',
        ]
    })
    exec_summary.to_excel(writer, sheet_name='Executive Summary', index=False)
    print(f"  ✓ Tab: Executive Summary")

    # Tab 2: Industry by BDC (Company Counts)
    industry_counts_export = industry_counts.copy()
    industry_counts_export.to_excel(writer, sheet_name='Industry Counts by BDC')
    print(f"  ✓ Tab: Industry Counts by BDC ({industry_counts.shape[0]} industries)")

    # Tab 3: Industry by BDC (Fair Value)
    industry_fv_export = industry_fv.copy()
    industry_fv_export.to_excel(writer, sheet_name='Industry Fair Value by BDC')
    print(f"  ✓ Tab: Industry Fair Value by BDC ({industry_fv.shape[0]} industries)")

    # Tab 4: Industry Overlap Analysis
    overlap_export = overlap_analysis.copy()
    overlap_export.to_excel(writer, sheet_name='Industry Overlap', index=False)
    print(f"  ✓ Tab: Industry Overlap ({len(overlap_analysis)} industries)")

    # Tab 5: ARCC Portfolio Detail
    arcc_df = df_analyzed[df_analyzed['bdc_name'] == 'Ares Capital Corporation'].copy()
    arcc_companies = arcc_df.groupby(['company_name', 'industry_sector']).agg({
        'fair_value': 'sum',
        'principal': 'sum',
        'cost': 'sum',
        'business_description': 'first'
    }).reset_index()
    arcc_companies['fair_value_millions'] = (arcc_companies['fair_value'] / 1e6).round(2)
    arcc_companies = arcc_companies.sort_values('fair_value', ascending=False)

    cols = ['company_name', 'industry_sector', 'business_description', 'fair_value_millions', 'fair_value']
    arcc_companies[cols].to_excel(writer, sheet_name='ARCC Portfolio', index=False)
    print(f"  ✓ Tab: ARCC Portfolio ({len(arcc_companies)} companies)")

    # Tab 6: MAIN Portfolio Detail
    main_df = df_analyzed[df_analyzed['bdc_name'] == 'Main Street Capital Corporation'].copy()
    main_companies = main_df.groupby(['company_name', 'industry_sector']).agg({
        'fair_value': 'sum',
        'principal': 'sum',
        'cost': 'sum',
        'business_description': 'first'
    }).reset_index()
    main_companies['fair_value_millions'] = (main_companies['fair_value'] / 1e6).round(2)
    main_companies = main_companies.sort_values('fair_value', ascending=False)

    main_companies[cols].to_excel(writer, sheet_name='MAIN Portfolio', index=False)
    print(f"  ✓ Tab: MAIN Portfolio ({len(main_companies)} companies)")

    # Tab 7: Concentration Metrics
    concentration_data = []

    for bdc in bdcs_with_data:
        bdc_df = df_analyzed[df_analyzed['bdc_name'] == bdc]
        bdc_companies = bdc_df.groupby(['company_name', 'industry_sector'])['fair_value'].sum().reset_index()

        # Exclude Unknown
        bdc_known = bdc_companies[bdc_companies['industry_sector'] != 'Unknown']
        total_fv = bdc_known['fair_value'].sum()

        # Industry shares
        industry_shares = bdc_known.groupby('industry_sector')['fair_value'].sum() / total_fv
        hhi = (industry_shares ** 2).sum() * 10000

        # Top industries
        top_industries = industry_shares.nlargest(5)

        for i, (industry, share) in enumerate(top_industries.items(), 1):
            concentration_data.append({
                'BDC': bdc,
                'Rank': i,
                'Industry': industry,
                'Fair Value ($B)': bdc_known[bdc_known['industry_sector'] == industry]['fair_value'].sum() / 1e9,
                'Portfolio %': share * 100,
                'Companies': len(bdc_known[bdc_known['industry_sector'] == industry])
            })

        # Add HHI row
        concentration_data.append({
            'BDC': bdc,
            'Rank': '',
            'Industry': 'HHI (Herfindahl-Hirschman Index)',
            'Fair Value ($B)': '',
            'Portfolio %': hhi,
            'Companies': ''
        })

    concentration_df = pd.DataFrame(concentration_data)
    concentration_df.to_excel(writer, sheet_name='Concentration Metrics', index=False)
    print(f"  ✓ Tab: Concentration Metrics")

    # Tab 8: Methodology & Limitations
    methodology_df = pd.DataFrame({
        'Section': [
            'Data Coverage',
            '',
            'BDCs Included',
            'BDCs Excluded',
            'Exclusion Reason',
            '',
            'Classification Method',
            'Industry Sectors',
            'Classification Algorithm',
            '',
            'Key Metrics',
            '',
            'Limitations',
            '',
            '',
            '',
            'HHI Interpretation',
            '',
            '',
            'Data Quality Notes',
            '',
            '',
        ],
        'Description': [
            '2 of 6 BDCs have business descriptions (41.6% coverage)',
            '',
            'Ares Capital Corporation (ARCC), Main Street Capital Corporation (MAIN)',
            'FS KKR Capital Corp, Prospect Capital, New Mountain Finance, Oaktree',
            'No business description field in parsed Schedule of Investments',
            '',
            'Keyword matching on business descriptions',
            '16 industry sectors plus "Other" and "Unknown" categories',
            'Rule-based classifier using industry-specific keywords',
            '',
            'HHI (Herfindahl-Hirschman Index): Measures portfolio concentration by industry',
            '',
            '1. Only 2 of 6 BDCs analyzed (60% of BDCs lack business descriptions)',
            '2. Results may not generalize to other BDCs',
            '3. Keyword-based classification may misclassify some companies',
            '',
            'HHI < 1500: Low concentration (diversified)',
            'HHI 1500-2500: Moderate concentration',
            'HHI > 2500: High concentration',
            '',
            'ARCC: 100% description coverage, 253 companies, $33.2B',
            'MAIN: 100% description coverage, 140 companies, $9.4B',
        ]
    })
    methodology_df.to_excel(writer, sheet_name='Methodology', index=False)
    print(f"  ✓ Tab: Methodology")

# Apply formatting
print("\nApplying formatting...")
wb = load_workbook(output_path)

header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True)
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    # Format headers (row 1)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = min(max_length + 2, 70)
        ws.column_dimensions[column_letter].width = adjusted_width

wb.save(output_path)
print("✓ Formatting applied")

print("\n" + "=" * 70)
print("REPORT GENERATION COMPLETE")
print("=" * 70)
print(f"\n✓ Excel report saved to: {output_path}")
print(f"\nTabs included:")
print(f"  1. Executive Summary")
print(f"  2. Industry Counts by BDC")
print(f"  3. Industry Fair Value by BDC")
print(f"  4. Industry Overlap")
print(f"  5. ARCC Portfolio (253 companies)")
print(f"  6. MAIN Portfolio (140 companies)")
print(f"  7. Concentration Metrics (HHI)")
print(f"  8. Methodology & Limitations")

print("\n" + "=" * 70)
print("KEY FINDINGS SUMMARY")
print("=" * 70)
print("\n1. INDUSTRY CONCENTRATION:")
print("   - ARCC: High concentration (HHI 2576)")
print("     • Software/Technology: 36.2% ($12.0B)")
print("     • Healthcare Services: 29.9% ($9.9B)")
print("     • Top 3 industries: 82.8% of portfolio")
print("\n   - MAIN: Moderate concentration (HHI 1702)")
print("     • Other/Mixed: 32.4% ($3.0B)")
print("     • Industrial/Manufacturing: 18.8% ($1.8B)")
print("     • Top 3 industries: 62.0% of portfolio")

print("\n2. INDUSTRY OVERLAP:")
print("   - Only 1 company overlaps (UserZoom in Software/Tech)")
print("   - 0.7% overlap rate in Software/Technology")
print("   - 0% overlap in all other industries")

print("\n3. PORTFOLIO DIFFERENTIATION:")
print("   - ARCC focuses on large-cap tech & healthcare")
print("   - MAIN is more diversified across industries")
print("   - Minimal competition even within same sectors")
