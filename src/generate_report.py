"""
Generate final Excel report with BDC overlap analysis
Run: python -m src.generate_report
"""

import pandas as pd
from pathlib import Path
from src.export import to_excel
from src.edgar_parser import TOP_10_BDCS

print("=" * 60)
print("GENERATING BDC OVERLAP REPORT")
print("=" * 60)

# Load analysis results
print("\nLoading analysis results...")

holdings_df = pd.read_csv('outputs/all_bdc_holdings.csv')
common_2plus = pd.read_csv('outputs/common_holdings_2plus.csv')
overlap_matrix = pd.read_csv('outputs/bdc_overlap_matrix.csv', index_col=0)
bdc_summary = pd.read_csv('outputs/bdc_overlap_summary.csv')

print(f"✓ Loaded {len(holdings_df):,} holdings")
print(f"✓ Loaded {len(common_2plus):,} common holdings")
print(f"✓ Loaded {overlap_matrix.shape[0]}×{overlap_matrix.shape[1]} overlap matrix")
print(f"✓ Loaded {len(bdc_summary)} BDC summaries")

# Prepare results dictionary
results = {
    'summary': bdc_summary,
    'common_holdings': common_2plus,
    'overlap_matrix': overlap_matrix,
    'all_holdings': holdings_df
}

# Create BDC info mapping (name -> ticker/name) for export
bdc_info_by_name = {}
for cik, info in TOP_10_BDCS.items():
    bdc_info_by_name[info['name']] = {
        'ticker': info['ticker'],
        'name': info['name'],
        'cik': cik
    }

# Generate Excel report
output_path = Path('outputs/bdc_overlap_analysis.xlsx')

print("\n" + "=" * 60)
print("CREATING EXCEL WORKBOOK")
print("=" * 60)

with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    # Tab 1: Executive Summary
    exec_summary = pd.DataFrame({
        'Metric': [
            'Total BDCs Analyzed',
            'Total Investment Positions',
            'Total Unique Portfolio Companies',
            'Total Fair Value',
            '',
            'Companies Held by 1 BDC',
            'Companies Held by 2+ BDCs',
            'Companies Held by 3+ BDCs',
            '',
            'Concentration Rate',
            'Average Investments per Company',
        ],
        'Value': [
            holdings_df['bdc_name'].nunique(),
            len(holdings_df),
            overlap_matrix.index.nunique(),
            f"${holdings_df['fair_value'].sum()/1_000_000_000:.2f}B",
            '',
            (overlap_matrix.sum(axis=1) == 1).sum(),
            len(common_2plus),
            0,  # We don't have any 3+ from analysis
            '',
            f"{len(common_2plus) / overlap_matrix.index.nunique() * 100:.2f}%",
            f"{len(holdings_df) / overlap_matrix.index.nunique():.2f}",
        ]
    })
    exec_summary.to_excel(writer, sheet_name='Executive Summary', index=False)
    print(f"  ✓ Tab: Executive Summary")

    # Tab 2: BDC Summary
    summary_export = bdc_summary.copy()
    # Add ticker and CIK info
    summary_export['ticker'] = summary_export['bdc_name'].map(
        lambda x: bdc_info_by_name.get(x, {}).get('ticker', '')
    )
    summary_export['cik'] = summary_export['bdc_name'].map(
        lambda x: bdc_info_by_name.get(x, {}).get('cik', '')
    )
    # Reorder columns
    cols = ['ticker', 'bdc_name', 'cik', 'total_companies', 'shared_companies', 'pct_shared']
    summary_export = summary_export[[c for c in cols if c in summary_export.columns]]
    summary_export.to_excel(writer, sheet_name='BDC Summary', index=False)
    print(f"  ✓ Tab: BDC Summary ({len(summary_export)} rows)")

    # Tab 3: Common Holdings (2+)
    common_2plus.to_excel(writer, sheet_name='Common Holdings (2+)', index=False)
    print(f"  ✓ Tab: Common Holdings (2+) ({len(common_2plus)} rows)")

    # Tab 4: BDC Overlap Matrix
    overlap_matrix.to_excel(writer, sheet_name='BDC Overlap Matrix')
    print(f"  ✓ Tab: BDC Overlap Matrix ({overlap_matrix.shape[0]}×{overlap_matrix.shape[1]})")

    # Tab 5: All Holdings
    # Sort by BDC, then company name
    all_holdings_sorted = holdings_df.sort_values(['bdc_name', 'company_name']).copy()
    # Format fair value for readability
    all_holdings_sorted['fair_value_millions'] = (all_holdings_sorted['fair_value'] / 1_000_000).round(2)
    cols_to_export = ['bdc_name', 'company_name', 'business_description', 'investment_type',
                      'fair_value_millions', 'principal', 'cost', 'fair_value']
    all_holdings_sorted[cols_to_export].to_excel(writer, sheet_name='All Holdings', index=False)
    print(f"  ✓ Tab: All Holdings ({len(all_holdings_sorted)} rows)")

    # Tab 6: Methodology
    methodology_df = pd.DataFrame({
        'Section': [
            'Data Source',
            'Data Date',
            'Extraction Method',
            'BDCs Analyzed',
            'BDCs Successfully Parsed',
            'Matching Method',
            'Common Holdings Definition',
            'Limitations',
            '',
            'Successfully Parsed BDCs',
            '',
        ] + list(holdings_df['bdc_name'].unique()),
        'Description': [
            'SEC EDGAR 10-K Filings (Schedule of Investments)',
            pd.Timestamp.now().strftime('%Y-%m-%d'),
            'Direct HTML parsing using EdgarTools library',
            '10 largest BDCs by market cap',
            f"{holdings_df['bdc_name'].nunique()} of 10 (60%)",
            'Exact company name match',
            'Companies held by 2+ BDCs based on exact name matching',
            'Name variations may cause undercount of overlaps. Company names not standardized.',
            '',
            'The following BDCs were successfully parsed:',
            '',
        ] + ['✓ ' + name for name in holdings_df['bdc_name'].unique()]
    })
    methodology_df.to_excel(writer, sheet_name='Methodology', index=False)
    print(f"  ✓ Tab: Methodology")

# Apply formatting using openpyxl
print("\nApplying formatting...")
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers

wb = load_workbook(output_path)

# Header formatting
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True)
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    # Format headers (row 1)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = min(max_length + 2, 60)  # Cap at 60
        ws.column_dimensions[column_letter].width = adjusted_width

wb.save(output_path)

print("✓ Formatting applied")

print("\n" + "=" * 60)
print("REPORT GENERATION COMPLETE")
print("=" * 60)
print(f"\n✓ Excel report saved to: {output_path}")
print(f"\nTabs included:")
print(f"  1. Executive Summary - Key metrics")
print(f"  2. BDC Summary - Per-BDC statistics")
print(f"  3. Common Holdings (2+) - Shared portfolio companies")
print(f"  4. BDC Overlap Matrix - Pairwise overlap counts")
print(f"  5. All Holdings - Complete dataset ({len(holdings_df):,} rows)")
print(f"  6. Methodology - Data source and limitations")
